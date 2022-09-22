# ======================================================
# ============= Travel Statement Generator =============
# ======================================================
#
# This program generates random travel times and
# calculates corresponding values and fees into
# a preformatted table
# 
# IMPORTANT ->
# 'empty.xlsx' must exist and has to be formatted properly!
# 'db.xlsx' must exist!
#
# only 4-city model (2 routes per day) applied

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import random, datetime, sys, decimal


# set number of days for months
DAYS_IN_MONTH = {1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31}

# if param is empty, show message
try:
    PARAM = sys.argv[1]
except IndexError:
    print("No parameter provided")
    print("Syntax: python generator.py [year]")
    sys.exit()


def checkParam():
    """ checks if provided parameter is correct """

    if not PARAM.isnumeric():
        print("Parameter is not integer")
        sys.exit()

    if int(PARAM) < 1970 or int(PARAM) > 2100:
        print("Year is out of range (1970-2100)")
        sys.exit()


def leapYearCheck():
    """ checks if provided year is leap-year """

    global DAYS_IN_MONTH
    try:
        if datetime.datetime(int(YEAR), 2, 29):
            DAYS_IN_MONTH[2] = 29
            return True

    except ValueError:
        DAYS_IN_MONTH[2] = 28
        return False


def generateAllYear():
    """ generate all 12 months """

    for month in range(1, 13):  # 12 months
        generateMonth(month)


def generateMonth(month):
    """ generate 1 month """

    ws = wb2["mesiac " + str(month)]  # select worksheet

    # fillSheet params: startRow, startColumn, startDate, month, numberOfDays, ws
    if month < 10:
        fillSheet(5, 1, datetime.datetime.strptime(YEAR + '-0' + str(month) + '-01', '%Y-%m-%d'), month, DAYS_IN_MONTH[month], ws)
    else:
        fillSheet(5, 1, datetime.datetime.strptime(YEAR + '-' + str(month) + '-01', '%Y-%m-%d'), month, DAYS_IN_MONTH[month], ws)


def getPetrol():
    """ count petrol consumption price per 1km """

    petrolPrice = random.randint(120, 135) / 100  # random cena za litr
    consumption = 22 / 100
    flatRateCompensation = .193  # pausalni nahrada
    consumptionPrice = petrolPrice * consumption + flatRateCompensation

    return [petrolPrice, consumptionPrice]


def randomTime(routeLength):
    """ generate random times and return start, end and total time """

    # convert routeLength minutes to a tuple (HOUR, MINUTE) and then to variables
    routeLengthHour = int(routeLength[:2])
    routeLengthMinute = int(routeLength[3:-3])

    # morning time generator
    randHour, randMinute = random.randint(0, 3), random.randint(0, 59)
    # salt = random.randint(0, 10)
    morningStartTime = datetime.datetime(1970, 1, 1, hour=6 + randHour, minute=randMinute)
    morningAdd = datetime.timedelta(hours=routeLengthHour, minutes=routeLengthMinute) # + salt)
    morningEndTime = morningStartTime + morningAdd
    
    # evening time generator
    randHour, randMinute = random.randint(0, 1), random.randint(27, 59)
    salt = random.randint(0, 10)
    eveningStartTime = datetime.datetime(1970, 1, 1, hour=19 + randHour, minute=randMinute)
    eveningAdd = datetime.timedelta(hours=routeLengthHour, minutes=routeLengthMinute) # + salt)
    eveningEndTime = eveningStartTime + eveningAdd

    # count total time (evening end - morning start)
    totalTime = str(eveningEndTime - morningStartTime)[:-3]

    # if total time is less than 10, add "0" to the beginning
    if int(totalTime[0]) < 10 and totalTime[1] == ":":
        totalTime = "0" + totalTime
    
    # convert times to strings and cut unwanted chars
    morningStartTime = str(morningStartTime)[10:-3]
    morningEndTime = str(morningEndTime)[10:-3]
    eveningStartTime = str(eveningStartTime)[10:-3]
    eveningEndTime = str(eveningEndTime)[10:-3]

    return [morningStartTime, morningEndTime, eveningStartTime, eveningEndTime, totalTime]


def dayRoute():
    """ generate random start city """

    dayRouteResult = []
    ws = wb1.active  # set input excel active
    randomStartCity = random.randint(1, 72)  # select random city combination from the base city

    cityFrom = ws['A'+str(randomStartCity)].value
    cityTo = ws['B'+str(randomStartCity)].value
    distance = ws['C'+str(randomStartCity)].value
    travelTime = ws['D'+str(randomStartCity)].value

    dayRouteResult.append(cityFrom)
    dayRouteResult.append(cityTo)
    dayRouteResult.append(str(distance))
    dayRouteResult.append(str(travelTime))

    ws = wb2.active  # set 2nd excel active

    return dayRouteResult


def repeat(currentMonth):
    """ repeat all process till "together" sum is under 1400 EUR """

    ws = wb2.active  # set 2nd excel active
    # generateDates params: startRow, startColumn, startDate, numberOfDays
    generateMonth(currentMonth)


def writeFooterValues(month, ws):
    """ fill footer with sum calculations """

    # write petrol column sum
    petrolValueList = []
    for row in range(5, (DAYS_IN_MONTH[month] * 4) + 4, 2):
        petrolValueList.append(ws.cell(row=row, column=8).value)
    sumOfValueList = sum(petrolValueList)
    ws.cell(row=129, column=8, value=sumOfValueList)

    # write diets column sum
    dietValueList = []
    for row in range(5, (DAYS_IN_MONTH[month] * 4) + 4, 2):
        dietValueList.append(ws.cell(row=row, column=9).value)
    sumOfValueList = sum(dietValueList)
    ws.cell(row=129, column=9, value=sumOfValueList)

    # write "together" fields sum
    togetherValueList = []
    for cell in range(8, 10):
        togetherValueList.append(float(ws.cell(row=129, column=cell).value))
    sumOfValueList = sum(togetherValueList)
    ws.cell(row=129, column=12, value=sumOfValueList)

    # repeat all calculations if total sum is over 1400 EUR
    if sumOfValueList > 1400:
        repeat(month)

    # write "overpay/underpay" field sum
    ws.cell(row=131, column=12, value=ws.cell(row=129, column=12).value)


def fillSheet(startRow, startColumn, startDate, month, numberOfDays, ws):
    """ fill the worksheet with all data """
    petrolPrice = getPetrol()  # get petrol price && consumption price (one for all month)


    # repeat it _numberOfDays_-times
    for day in range(numberOfDays):
        getStartRoute = dayRoute()  # get dayRoute function result [start city, end city, km, time]
        getRandomTime = randomTime(getStartRoute[3])  # get start hours, end hours and travel time
                                                      # [start time, end time, start time2, end time2, total time]

        # keep generating new time until it's less than 12:00
        while str(getRandomTime[4][0]) == "1" and int(getRandomTime[4][1]) > 1:
            getStartRoute = dayRoute()  # if 12 or more, do again!
            getRandomTime = randomTime(getStartRoute[3])

        # set diets
        # if 5-12h, set to 5.1
        if int(getRandomTime[4][:2]) >= 5 and int(getRandomTime[4][:2]) < 12:
            diets = round(decimal.Decimal(5.1), 2)
        #if 12-18h, set to 7.6
        if int(getRandomTime[4][:2]) >= 12 and int(getRandomTime[4][:2]) < 18:
            diets = round(decimal.Decimal(7.6), 2)
        # if 18h and more, set to 11.6
        if int(getRandomTime[4][:2]) >= 18:
            diets = round(decimal.Decimal(11.6), 2)

        for event in range(4):
            # way there ->
            if event == 0:
                startDateStr = startDate.strftime('%d.%m.%Y')  # starting date (usualy 1.1.2020)
                ws.cell(row=startRow, column=startColumn, value=startDateStr)  # fill 1st cell with date
                ws.cell(row=startRow, column=startColumn+1, value="odchod")  # fill odchod
                ws.cell(row=startRow, column=startColumn+2, value=getStartRoute[0])  # fill start city
                ws.cell(row=startRow, column=startColumn+3, value=getRandomTime[0]).alignment = Alignment(vertical="center", horizontal="center")  # morning start time
                ws.cell(row=startRow, column=startColumn+4, value="AUS").alignment = Alignment(vertical="center", horizontal="center")  # set to AUS
                ws.cell(row=startRow, column=startColumn+5, value=getStartRoute[2]).alignment = Alignment(vertical="center", horizontal="center")  # km

                # ws.cell(row=1, column=13, value=petrolPrice[0]) # test!!! write petrol price to table
                petrolPriceKm = petrolPrice[1] * float(getStartRoute[2])
                petrolPriceKm = float(round(decimal.Decimal(petrolPriceKm), 2))

                ws.cell(row=startRow, column=startColumn+7, value=petrolPriceKm).alignment = Alignment(vertical="center", horizontal="center")  # petrol price
                ws.cell(row=startRow, column=startColumn+8, value=diets).alignment = Alignment(vertical="center", horizontal="center")  # diets
                ws.cell(row=startRow, column=startColumn+11, value=petrolPriceKm+float(diets)).alignment = Alignment(vertical="center", horizontal="center")  # together
            if event == 1:
                ws.cell(row=startRow, column=startColumn+1, value="príchod")  # fill prichod
                ws.cell(row=startRow, column=startColumn+2, value=getStartRoute[1])  # fill destination city
                ws.cell(row=startRow, column=startColumn+3, value=getRandomTime[1]).alignment = Alignment(vertical="center", horizontal="center")  # morning end time

            # way back <-
            if event == 2:
                ws.cell(row=startRow, column=startColumn+1, value="odchod")  # fill odchod
                ws.cell(row=startRow, column=startColumn+2, value=getStartRoute[1])  # fill destination city
                ws.cell(row=startRow, column=startColumn+3, value=getRandomTime[2]).alignment = Alignment(vertical="center", horizontal="center")  # evening start time
                ws.cell(row=startRow, column=startColumn+4, value="AUS").alignment = Alignment(vertical="center", horizontal="center")
                ws.cell(row=startRow, column=startColumn+5, value=getStartRoute[2]).alignment = Alignment(vertical="center", horizontal="center")  # km
                ws.cell(row=startRow, column=startColumn+7, value=petrolPriceKm).alignment = Alignment(vertical="center", horizontal="center")  # petrol price
                ws.cell(row=startRow, column=startColumn+8, value=diets).alignment = Alignment(vertical="center", horizontal="center")  # diets
                ws.cell(row=startRow, column=startColumn+11, value=petrolPriceKm+float(diets)).alignment = Alignment(vertical="center", horizontal="center")  # together
            if event == 3:
                ws.cell(row=startRow, column=startColumn+1, value="príchod")  # fill prichod
                ws.cell(row=startRow, column=startColumn+2, value=getStartRoute[0])  # fill start city
                ws.cell(row=startRow, column=startColumn+3, value=getRandomTime[3]).alignment = Alignment(vertical="center", horizontal="center")  # evening end time

            startRow += 1
        
        # increment day by 1
        startDate = startDate + datetime.timedelta(days=1)

    writeFooterValues(month, ws)

# main
if __name__ == '__main__':
    checkParam()
    YEAR = PARAM
    isLeap = leapYearCheck()

    print("Loading input files")
    wb1 = load_workbook("db.xlsx")
    wb2 = load_workbook("empty.xlsx")
    ws = wb2.active  # set 2nd excel active

    print("Generating worksheets for year", YEAR, end=" ")
    if isLeap:
        print("(leap-year)")
    else:
        print()

    generateAllYear()

    print("Saving output file")
    wb2.save("output.xlsx")